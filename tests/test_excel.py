"""
Unit tests for Excel output generation functions
"""

import unittest
import pandas as pd
import os
import tempfile
from openpyxl import load_workbook
import sys
from pathlib import Path

# Add parent directory to path to import package
sys.path.append(str(Path(__file__).parent.parent))

from cashflow_tracker.core.schema import TransactionSchema, CategorySchema
from cashflow_tracker.output.excel import (
    create_excel_workbook, populate_transaction_sheet,
    populate_category_sheet, populate_summary_sheet,
    create_charts
)


class TestExcel(unittest.TestCase):
    """Test Excel file generation functions"""

    def setUp(self):
        """Set up test fixtures"""
        # Create test transaction data
        self.transaction_data = pd.DataFrame({
            TransactionSchema.DATE: pd.to_datetime(['2025-04-01', '2025-04-15']),
            TransactionSchema.DESCRIPTION: ['Salary', 'Rent'],
            TransactionSchema.AMOUNT: [3000.0, 1200.0],
            TransactionSchema.TYPE: ['Income', 'Expense'],
            TransactionSchema.CATEGORY: ['Income', 'Housing'],
            TransactionSchema.SUBCATEGORY: ['Salary', 'Rent/Mortgage'],
            TransactionSchema.PRODUCER: ['Employer', 'Landlord'],
            TransactionSchema.PAYMENT_METHOD: ['Direct Deposit', 'Bank Transfer'],
            TransactionSchema.NOTES: [None, None]
        })

        # Create test category data
        self.category_data = pd.DataFrame({
            CategorySchema.MAIN_CATEGORY: ['Income', 'Housing', 'Food'],
            CategorySchema.SUBCATEGORY: [
                ['Salary', 'Freelance'],
                ['Rent/Mortgage', 'Utilities'],
                ['Groceries', 'Restaurants']
            ],
            CategorySchema.DESCRIPTION: [
                'Income sources',
                'Housing expenses',
                'Food expenses'
            ],
            CategorySchema.BUDGET: [0.0, 1500.0, 500.0]
        })

        # Create aggregated data
        self.type_summary = {'Income': 3000.0, 'Expense': 1200.0}

        self.category_summary = pd.DataFrame({
            TransactionSchema.TYPE: ['Income', 'Expense'],
            TransactionSchema.CATEGORY: ['Income', 'Housing'],
            TransactionSchema.AMOUNT: [3000.0, 1200.0]
        })

        self.cash_allocation = {
            'Spending': 100.0,  # 100% of expenses are spending
            'Saving': 0.0,
            'Investing': 0.0
        }

        self.monthly_summary = pd.DataFrame({
            'Month': ['2025-04', '2025-04'],
            TransactionSchema.TYPE: ['Income', 'Expense'],
            TransactionSchema.AMOUNT: [3000.0, 1200.0]
        })

        # Create temporary file
        self.temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.excel_filename = self.temp_file.name
        self.temp_file.close()

    def tearDown(self):
        """Tear down test fixtures"""
        os.unlink(self.excel_filename)

    def test_create_excel_workbook(self):
        """Test Excel workbook creation"""
        wb = create_excel_workbook()

        # Check that the workbook has the expected sheets
        sheet_names = wb.sheetnames
        self.assertIn("Transaction Log", sheet_names)
        self.assertIn("Categories", sheet_names)
        self.assertIn("Summary", sheet_names)
        self.assertIn("Charts", sheet_names)

        # Save workbook to test file
        wb.save(self.excel_filename)

        # Load the workbook and check again
        loaded_wb = load_workbook(self.excel_filename)
        loaded_sheet_names = loaded_wb.sheetnames
        self.assertIn("Transaction Log", loaded_sheet_names)
        self.assertIn("Categories", loaded_sheet_names)
        self.assertIn("Summary", loaded_sheet_names)
        self.assertIn("Charts", loaded_sheet_names)

    def test_populate_transaction_sheet(self):
        """Test transaction sheet population"""
        wb = create_excel_workbook()
        populate_transaction_sheet(wb, self.transaction_data)

        # Save workbook to test file
        wb.save(self.excel_filename)

        # Load the workbook and check the transaction sheet
        loaded_wb = load_workbook(self.excel_filename)
        ws = loaded_wb["Transaction Log"]

        # Check headers
        for col, header in enumerate(TransactionSchema.get_columns(), 1):
            self.assertEqual(ws.cell(row=1, column=col).value, header)

        # Check data in first row
        self.assertEqual(ws.cell(row=2, column=2).value, "Salary")  # Description
        self.assertEqual(ws.cell(row=2, column=3).value, 3000.0)  # Amount
        self.assertEqual(ws.cell(row=2, column=4).value, "Income")  # Type

        # Check data in second row
        self.assertEqual(ws.cell(row=3, column=2).value, "Rent")  # Description
        self.assertEqual(ws.cell(row=3, column=3).value, 1200.0)  # Amount
        self.assertEqual(ws.cell(row=3, column=4).value, "Expense")  # Type

    def test_populate_category_sheet(self):
        """Test category sheet population"""
        wb = create_excel_workbook()
        populate_category_sheet(wb, self.category_data)

        # Save workbook to test file
        wb.save(self.excel_filename)

        # Load the workbook and check the category sheet
        loaded_wb = load_workbook(self.excel_filename)
        ws = loaded_wb["Categories"]

        # Check headers
        for col, header in enumerate(CategorySchema.get_columns(), 1):
            self.assertEqual(ws.cell(row=1, column=col).value, header)

        # Check data
        self.assertEqual(ws.cell(row=2, column=1).value, "Income")  # Main Category
        self.assertEqual(ws.cell(row=3, column=1).value, "Housing")  # Main Category
        self.assertEqual(ws.cell(row=3, column=4).value, 1500.0)  # Budget Amount

    def test_populate_summary_sheet(self):
        """Test summary sheet population"""
        wb = create_excel_workbook()
        populate_summary_sheet(
            wb,
            self.type_summary,
            self.category_summary,
            self.cash_allocation
        )

        # Save workbook to test file
        wb.save(self.excel_filename)

        # Load the workbook and check the summary sheet
        loaded_wb = load_workbook(self.excel_filename)
        ws = loaded_wb["Summary"]

        # Check overall summary
        self.assertEqual(ws['A3'].value, "Total Income:")
        self.assertEqual(ws['B3'].value, 3000.0)
        self.assertEqual(ws['A4'].value, "Total Expenses:")
        self.assertEqual(ws['B4'].value, 1200.0)
        self.assertEqual(ws['A5'].value, "Net Cashflow:")
        self.assertEqual(ws['B5'].value, 1800.0)  # 3000 - 1200

    def test_create_charts(self):
        """Test chart creation"""
        wb = create_excel_workbook()

        # First populate the transaction sheet (required for charts)
        populate_transaction_sheet(wb, self.transaction_data)

        # Create charts
        create_charts(
            wb,
            self.category_summary,
            self.cash_allocation,
            self.monthly_summary
        )

        # Save workbook to test file
        wb.save(self.excel_filename)

        # Load the workbook and check the charts sheet
        loaded_wb = load_workbook(self.excel_filename)
        ws = loaded_wb["Charts"]

        # Check that the chart sheet has at least some data
        self.assertTrue(ws['A1'].value is not None)

        # Note: Testing the actual charts is difficult because openpyxl loads
        # charts differently than they are created. We'll just check that
        # the sheet exists and has some data.


if __name__ == '__main__':
    unittest.main()