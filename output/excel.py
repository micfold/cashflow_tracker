"""
Excel output generation functions for the Cashflow Tracker
"""

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
from typing import Dict, Optional
from core.schema import TransactionSchema, CategorySchema


def create_excel_workbook() -> Workbook:
    """
    Create an Excel workbook with standard sheets

    Returns:
        Openpyxl Workbook
    """
    wb = Workbook()

    # Rename default sheet
    ws = wb.active
    ws.title = "Transaction Log"

    # Create additional sheets
    wb.create_sheet("Categories")
    wb.create_sheet("Summary")
    wb.create_sheet("Charts")

    return wb


def populate_transaction_sheet(wb: Workbook,
                               transaction_data: pd.DataFrame) -> None:
    """
    Populate the transaction log sheet with data

    Args:
        wb: Openpyxl Workbook
        transaction_data: DataFrame with transaction data
    """
    ws = wb["Transaction Log"]

    # Convert data to Excel format
    for r_idx, row in enumerate(dataframe_to_rows(transaction_data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Format headers
    for col in range(1, len(TransactionSchema.get_columns()) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Format date column
    for row in range(2, len(transaction_data) + 2):
        cell = ws.cell(row=row, column=1)
        if cell.value:
            cell.number_format = 'YYYY-MM-DD'

    # Format amount column
    for row in range(2, len(transaction_data) + 2):
        cell = ws.cell(row=row, column=3)
        if cell.value:
            cell.number_format = '#,##0.00'

    # Add conditional formatting for transaction types
    income_rule = CellIsRule(operator='equal', formula=['"Income"'],
                             stopIfTrue=True, fill=PatternFill(bgColor="C6EFCE"))
    expense_rule = CellIsRule(operator='equal', formula=['"Expense"'],
                              stopIfTrue=True, fill=PatternFill(bgColor="FFCCCC"))

    type_col = TransactionSchema.get_columns().index(TransactionSchema.TYPE) + 1
    if len(transaction_data) > 0:
        ws.conditional_formatting.add(f"{chr(64 + type_col)}2:{chr(64 + type_col)}{len(transaction_data) + 1}",
                                  income_rule)
        ws.conditional_formatting.add(f"{chr(64 + type_col)}2:{chr(64 + type_col)}{len(transaction_data) + 1}",
                                  expense_rule)

    # Auto-fit columns
    for col in range(1, len(TransactionSchema.get_columns()) + 1):
        column_letter = chr(64 + col)
        ws.column_dimensions[column_letter].auto_size = True


def populate_category_sheet(wb: Workbook,
                            category_data: pd.DataFrame) -> None:
    """
    Populate the categories sheet with data

    Args:
        wb: Openpyxl Workbook
        category_data: DataFrame with category data
    """
    ws = wb["Categories"]

    # Create a copy of the dataframe to avoid modifying the original
    excel_data = category_data.copy()

    # Convert lists to comma-separated strings
    for col in excel_data.columns:
        if excel_data[col].apply(lambda x: isinstance(x, list)).any():
            excel_data[col] = excel_data[col].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)

    # Write the data (including header)
    for r_idx, row in enumerate(dataframe_to_rows(excel_data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Format headers
    for col in range(1, len(CategorySchema.get_columns()) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Format budget column (assuming column 4 is 'Budget Amount')
    for row in range(2, len(category_data) + 2):
        cell = ws.cell(row=row, column=4)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'

    # Auto-fit columns (rough approach since openpyxl doesn't auto-fit natively)
    for col in range(1, len(CategorySchema.get_columns()) + 1):
        column_letter = chr(64 + col)
        ws.column_dimensions[column_letter].width = 20  # Manually set a decent width



def populate_summary_sheet(wb: Workbook,
                           type_summary: Dict[str, float],
                           category_summary: pd.DataFrame,
                           cash_allocation: Dict[str, float],
                           budget_comparison: Optional[pd.DataFrame] = None) -> None:
    """
    Populate the summary sheet with aggregated data

    Args:
        wb: Openpyxl Workbook
        type_summary: Dict with income/expense totals
        category_summary: DataFrame with category totals
        cash_allocation: Dict with spending/saving/investing percentages
        budget_comparison: DataFrame with budget comparison
    """
    ws = wb["Summary"]

    # 1. Overall Summary Section
    ws['A1'] = "OVERALL SUMMARY"
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = "Total Income:"
    ws['B3'] = type_summary.get('Income', 0)
    ws['B3'].number_format = '#,##0.00'

    ws['A4'] = "Total Expenses:"
    ws['B4'] = type_summary.get('Expense', 0)
    ws['B4'].number_format = '#,##0.00'

    ws['A5'] = "Net Cashflow:"
    ws['B5'] = type_summary.get('Income', 0) - type_summary.get('Expense', 0)
    ws['B5'].number_format = '#,##0.00'

    # Highlight positive/negative cashflow
    if (type_summary.get('Income', 0) - type_summary.get('Expense', 0)) >= 0:
        ws['B5'].fill = PatternFill(fgColor="C6EFCE", fill_type="solid")
    else:
        ws['B5'].fill = PatternFill(fgColor="FFCCCC", fill_type="solid")

    # 2. Category Summary Section
    ws['A7'] = "CATEGORY SUMMARY"
    ws['A7'].font = Font(bold=True, size=14)

    ws['A9'] = "Category"
    ws['B9'] = "Amount"
    ws['C9'] = "% of Total"

    expense_categories = category_summary[category_summary[TransactionSchema.TYPE] == 'Expense']
    total_expense = type_summary.get('Expense', 0)

    for i, row in enumerate(expense_categories.itertuples(), 10):
        ws[f'A{i}'] = row.Category
        ws[f'B{i}'] = row.Amount
        ws[f'B{i}'].number_format = '#,##0.00'

        if total_expense > 0:
            ws[f'C{i}'] = row.Amount / total_expense
            ws[f'C{i}'].number_format = '0.0%'

    # 3. Cash Allocation Section
    allocation_start_row = 10 + len(expense_categories) + 2

    ws[f'A{allocation_start_row}'] = "CASH ALLOCATION"
    ws[f'A{allocation_start_row}'].font = Font(bold=True, size=14)

    ws[f'A{allocation_start_row + 2}'] = "Spending:"
    ws[f'B{allocation_start_row + 2}'] = cash_allocation.get('Spending', 0) / 100
    ws[f'B{allocation_start_row + 2}'].number_format = '0.0%'

    ws[f'A{allocation_start_row + 3}'] = "Saving:"
    ws[f'B{allocation_start_row + 3}'] = cash_allocation.get('Saving', 0) / 100
    ws[f'B{allocation_start_row + 3}'].number_format = '0.0%'

    ws[f'A{allocation_start_row + 4}'] = "Investing:"
    ws[f'B{allocation_start_row + 4}'] = cash_allocation.get('Investing', 0) / 100
    ws[f'B{allocation_start_row + 4}'].number_format = '0.0%'

    # 4. Budget Comparison (if provided)
    if budget_comparison is not None and not budget_comparison.empty:
        budget_start_row = allocation_start_row + 6

        ws[f'A{budget_start_row}'] = "BUDGET COMPARISON"
        ws[f'A{budget_start_row}'].font = Font(bold=True, size=14)

        ws[f'A{budget_start_row + 2}'] = "Category"
        ws[f'B{budget_start_row + 2}'] = "Actual"
        ws[f'C{budget_start_row + 2}'] = "Budget"
        ws[f'D{budget_start_row + 2}'] = "% Used"
        ws[f'E{budget_start_row + 2}'] = "Difference"

        for i, row in enumerate(budget_comparison.itertuples(), budget_start_row + 3):
            ws[f'A{i}'] = row.Category
            ws[f'B{i}'] = row.Amount
            ws[f'B{i}'].number_format = '#,##0.00'

            ws[f'C{i}'] = getattr(row, CategorySchema.BUDGET)
            ws[f'C{i}'].number_format = '#,##0.00'

            ws[f'D{i}'] = getattr(row, 'Budget Used (%)') / 100
            ws[f'D{i}'].number_format = '0.0%'

            ws[f'E{i}'] = getattr(row, 'Difference')
            ws[f'E{i}'].number_format = '#,##0.00'

            # Highlight over/under budget
            if getattr(row, 'Budget Used (%)') > 100:
                ws[f'D{i}'].fill = PatternFill(fgColor="FFCCCC", fill_type="solid")  # Red for over budget
            elif getattr(row, 'Budget Used (%)') > 90:
                ws[f'D{i}'].fill = PatternFill(fgColor="FFEB9C", fill_type="solid")  # Yellow for near budget
            else:
                ws[f'D{i}'].fill = PatternFill(fgColor="C6EFCE", fill_type="solid")  # Green for under budget


def create_charts(wb: Workbook,
                  category_summary: pd.DataFrame,
                  cash_allocation: Dict[str, float],
                  monthly_summary: pd.DataFrame) -> None:
    """
    Create charts for visualizing financial data

    Args:
        wb: Openpyxl Workbook
        category_summary: DataFrame with category totals
        cash_allocation: Dict with spending/saving/investing percentages
        monthly_summary: DataFrame with monthly totals
    """
    ws = wb["Charts"]

    # 1. Spending by Category Pie Chart
    pie_chart1 = PieChart()
    pie_chart1.title = "Spending by Category"

    # Filter for expense categories only
    expense_data = category_summary[category_summary[TransactionSchema.TYPE] == 'Expense']

    # Write category data to chart sheet for reference
    ws['A1'] = "Category"
    ws['B1'] = "Amount"

    for i, row in enumerate(expense_data.itertuples(), 2):
        ws[f'A{i}'] = row.Category
        ws[f'B{i}'] = row.Amount

    # Create references for chart data
    labels = Reference(ws, min_col=1, min_row=2, max_row=1 + len(expense_data))
    data = Reference(ws, min_col=2, min_row=2, max_row=1 + len(expense_data))

    # Add data to chart
    pie_chart1.add_data(data)
    pie_chart1.set_categories(labels)

    # Add data labels
    pie_chart1.dataLabels = DataLabelList()
    pie_chart1.dataLabels.showCatName = True
    pie_chart1.dataLabels.showPercent = True

    # Add chart to worksheet
    ws.add_chart(pie_chart1, "D1")

    # 2. Cash Allocation Pie Chart
    pie_chart2 = PieChart()
    pie_chart2.title = "Cash Allocation"

    # Write allocation data to chart sheet
    ws['A15'] = "Allocation"
    ws['B15'] = "Percentage"

    allocation_labels = ["Spending", "Saving", "Investing"]
    for i, label in enumerate(allocation_labels, 16):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = cash_allocation.get(label, 0)

    # Create references for chart data
    labels = Reference(ws, min_col=1, min_row=16, max_row=18)
    data = Reference(ws, min_col=2, min_row=16, max_row=18)

    # Add data to chart
    pie_chart2.add_data(data)
    pie_chart2.set_categories(labels)

    # Add data labels
    pie_chart2.dataLabels = DataLabelList()
    pie_chart2.dataLabels.showCatName = True
    pie_chart2.dataLabels.showPercent = True

    # Add chart to worksheet
    ws.add_chart(pie_chart2, "D15")

    # 3. Monthly Income vs Expenses Bar Chart (if data is available)
    if not monthly_summary.empty:
        # Pivot the data for easier plotting
        monthly_pivot = pd.pivot_table(
            monthly_summary,
            index='Month',
            columns=TransactionSchema.TYPE,
            values=TransactionSchema.AMOUNT,
            aggfunc='sum'
        ).fillna(0)

        # Sort by month
        monthly_pivot = monthly_pivot.sort_index()

        # Write monthly data to chart sheet
        ws['A30'] = "Month"
        ws['B30'] = "Income"
        ws['C30'] = "Expense"

        for i, (month, row) in enumerate(monthly_pivot.iterrows(), 31):
            ws[f'A{i}'] = month
            ws[f'B{i}'] = row.get('Income', 0)
            ws[f'C{i}'] = row.get('Expense', 0)

        # Create bar chart
        bar_chart = BarChart()
        bar_chart.title = "Monthly Income vs Expenses"
        bar_chart.style = 10
        bar_chart.x_axis.title = "Month"
        bar_chart.y_axis.title = "Amount"

        # Add data series
        cats = Reference(ws, min_col=1, min_row=31, max_row=30 + len(monthly_pivot))
        income_data = Reference(ws, min_col=2, min_row=30, max_row=30 + len(monthly_pivot))
        expense_data = Reference(ws, min_col=3, min_row=30, max_row=30 + len(monthly_pivot))

        bar_chart.add_data(income_data, titles_from_data=True)
        bar_chart.add_data(expense_data, titles_from_data=True)
        bar_chart.set_categories(cats)

        # Add to worksheet
        ws.add_chart(bar_chart, "D30")